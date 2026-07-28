from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def export_excel(ip, excel_data):

    wb = Workbook()
    ws = wb.active
    ws.title = "Huawei Discovery"

    # ===== Report Header =====
    ws.merge_cells("A1:C1")
    ws["A1"] = "Huawei Discovery Tool V2.1"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:C2")
    ws["A2"] = "Developed by Sun Technologies"
    ws["A2"].font = Font(size=12, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    # ===== Switch Info =====
    ws["A4"] = "Switch IP"
    ws["B4"] = ip
    
    ws["A5"] = "Model"
    ws["B5"] = model

    ws["A6"] = "Serial Number"
    ws["B6"] = serial

    ws["A7"] = "Software Version"
    ws["B7"] = software

    ws["A8"] = "Hardware Version"
    ws["B8"] = hardware

    # ===== Table Header =====
    ws["A10"] = "Port Name"
    ws["B10"] = "ifIndex"
    ws["C10"] = "entPhysicalIndex"

    for cell in ["A10", "B10", "C10"]:
        ws[cell].font = Font(bold=True)

    # ===== Data =====
    row = 11

    for data in excel_data:
        ws.cell(row=row, column=1).value = data[0]
        ws.cell(row=row, column=2).value = data[1]
        ws.cell(row=row, column=3).value = data[2]
        row += 1

    wb.save(f"HuaweiDiscovery_{ip}.xlsx")
